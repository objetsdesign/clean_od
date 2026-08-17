# -*- coding: utf-8 -*-
import base64
import io
import logging

from odoo import fields, models
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:  # pragma: no cover
    openpyxl = None


# ---------------------------------------------------------------------------
# Column header synonyms (lower-cased, stripped) -> internal key.
# Several columns in the original file are not always at the same index
# (an extra blank row, a missing column, ...) so we always locate columns
# by reading the header row instead of relying on fixed positions.
# ---------------------------------------------------------------------------
QUOTE_HEADERS = {
    'n° devis': 'name',
    'n° commande lié': 'order_ref',
    'date devis': 'date_devis',
    'client': 'client',
    'secteur': 'secteur',
    'contact': 'contact',
    'description produit(s)': 'description',
    'qté': 'qty',
    'montant ht (€)': 'amount',
    'statut devis': 'state',
    'date relance 1': 'date_relance_1',
    'date relance 2': 'date_relance_2',
    'date relance 3': 'date_relance_3',
    'réponse client': 'reponse_client',
    'motif refus / blocage': 'motif_refus',
    'proba. conversion %': 'proba_conversion',
    'action suivante': 'next_action',
    'notes': 'notes',
}

ORDER_HEADERS = {
    'n° commande': 'name',
    'n° devis lié': 'quote_ref',
    'date commande': 'date_commande',
    'client': 'client',
    'contact': 'contact',
    'description produit(s)': 'description',
    'qté': 'qty',
    'montant ht (€)': 'amount_ht',
    'coût (€)': 'cost',
    'acompte reçu': 'deposit_received',
    'solde à recevoir': 'balance_due',
    'statut commande': 'state',
    'source production': 'production_source',
    'date prod. prévue': 'date_prod_prevue',
    'date prod. réelle': 'date_prod_reelle',
    'date expédition': 'date_expedition',
    'transporteur': 'carrier',
    'frais logistiques': 'logistics_fees',
    'n° tracking': 'tracking_number',
    'date livraison prévue': 'date_livraison_prevue',
    'date livraison réelle': 'date_livraison_reelle',
    'notes / incidents': 'notes_incidents',
}

QUOTE_STATE_MAP = {
    'en cours': 'en_cours',
    'envoyé': 'envoye',
    'envoye': 'envoye',
    'relancé': 'relance',
    'relance': 'relance',
    'accepté': 'accepte',
    'accepte': 'accepte',
    'refusé': 'refuse',
    'refuse': 'refuse',
    'expiré': 'expire',
    'expire': 'expire',
    'en attente client': 'attente_client',
}

ORDER_STATE_MAP = {
    'confirmée': 'confirmee',
    'confirmee': 'confirmee',
    'en production': 'en_production',
    'contrôle qualité': 'controle_qualite',
    'controle qualite': 'controle_qualite',
    'prête à expédier': 'prete_expedier',
    'prete a expedier': 'prete_expedier',
    'expédiée': 'expediee',
    'expediee': 'expediee',
    'livrée': 'livree',
    'livree': 'livree',
    'litige': 'litige',
    'annulée': 'annulee',
    'annulee': 'annulee',
}

PRODUCTION_SOURCE_MAP = {
    'tunisie vri': 'tunisie_vri',
    'tunisie bsi': 'tunisie_bsi',
    'chine': 'chine',
    'stock europe': 'stock_europe',
    'multi-source': 'multi_source',
    'multi source': 'multi_source',
    'vietnam': 'vietnam',
    'inde': 'inde',
    'bangladesh': 'bangladesh',
    'multi-source asie': 'multi_source_asie',
    'multi source asie': 'multi_source_asie',
}

ACTIVITY_SHEET_MAP = [
    ('b2b', 'b2b'),
    ('fund', 'fund'),
    ('asie', 'asie'),
]


def _norm(value):
    return (str(value).strip().lower()) if value not in (None, '') else ''


def _to_str(value):
    if value is None:
        return False
    if isinstance(value, float) and value.is_integer():
        value = int(value)
    return str(value).strip() or False


def _to_float(value):
    if value in (None, ''):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _to_date(value):
    if not value:
        return False
    if hasattr(value, 'date'):
        return value.date()
    return False


class B2faImportWizard(models.TransientModel):
    _name = 'b2fa.import.wizard'
    _description = "Import Devis & Commandes depuis Excel"

    file = fields.Binary(string="Fichier Excel (.xlsx)", required=True)
    filename = fields.Char(string="Nom du fichier")
    result_log = fields.Text(string="Résultat de l'import", readonly=True)

    # -- helpers -----------------------------------------------------
    def _find_header_row(self, ws, known_headers):
        """Return (row_index, {col_index: internal_key}) for the header row."""
        for r in range(1, min(ws.max_row, 15) + 1):
            col_map = {}
            for c in range(1, ws.max_column + 1):
                key = _norm(ws.cell(row=r, column=c).value)
                if key in known_headers:
                    col_map[c] = known_headers[key]
            if col_map:
                return r, col_map
        return None, {}

    def _row_is_empty(self, row_vals):
        # Ignore rows where only "notes"/"action suivante"/"proba" style stray
        # cells are filled but there is no real record (client/amount/description).
        meaningful_keys = ('client', 'amount', 'amount_ht', 'description', 'name', 'order_ref', 'quote_ref')
        return not any(row_vals.get(k) not in (None, '', False) for k in meaningful_keys)

    def _map_state(self, raw, mapping):
        key = _norm(raw)
        return mapping.get(key, False)

    def _map_production_source(self, raw):
        key = _norm(raw)
        return PRODUCTION_SOURCE_MAP.get(key, False)

    def _map_proba(self, raw):
        if raw in (None, ''):
            return False
        text = str(raw).replace('%', '').strip()
        allowed = ('10', '20', '30', '50', '70', '90', '100')
        return text if text in allowed else False

    # -- main entry point ---------------------------------------------
    def action_import(self):
        self.ensure_one()
        if not openpyxl:
            raise UserError(
                "La librairie Python 'openpyxl' n'est pas disponible sur ce serveur Odoo. "
                "Demandez à votre administrateur système de l'installer (pip install openpyxl)."
            )
        if not self.file:
            raise UserError("Merci de sélectionner un fichier Excel (.xlsx) à importer.")

        try:
            content = base64.b64decode(self.file)
            wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
        except Exception as exc:
            raise UserError("Impossible de lire ce fichier Excel : %s" % exc)

        Quote = self.env['b2fa.quote']
        Order = self.env['b2fa.order']

        log_lines = []
        quotes_created = 0
        quotes_skipped = 0
        orders_created = 0
        orders_skipped = 0

        # Pass 1: import all "DEVIS ..." sheets first so orders can be linked afterwards.
        for sheet_name in wb.sheetnames:
            norm_name = _norm(sheet_name)
            if not norm_name.startswith('devis'):
                continue
            activity = self._detect_activity(norm_name)
            if not activity:
                continue
            SheetQuote, _SheetOrder = self._models_for_activity(activity)
            ws = wb[sheet_name]
            header_row, col_map = self._find_header_row(ws, QUOTE_HEADERS)
            if not header_row:
                log_lines.append("⚠ Onglet '%s' : en-têtes non reconnus, ignoré." % sheet_name)
                continue

            created, skipped = self._import_quote_sheet(ws, header_row, col_map, activity, SheetQuote)
            quotes_created += created
            quotes_skipped += skipped
            log_lines.append("📝 %s : %s devis importé(s), %s ignoré(s) (déjà présents ou vides)."
                              % (sheet_name.strip(), created, skipped))

        # Pass 2: import all "COMMANDES ..." sheets and try to link to quotes.
        for sheet_name in wb.sheetnames:
            norm_name = _norm(sheet_name)
            if not norm_name.startswith('commandes'):
                continue
            activity = self._detect_activity(norm_name)
            if not activity:
                continue
            SheetQuote, SheetOrder = self._models_for_activity(activity)
            ws = wb[sheet_name]
            header_row, col_map = self._find_header_row(ws, ORDER_HEADERS)
            if not header_row:
                log_lines.append("⚠ Onglet '%s' : en-têtes non reconnus, ignoré." % sheet_name)
                continue

            created, skipped = self._import_order_sheet(ws, header_row, col_map, activity, SheetOrder, SheetQuote)
            orders_created += created
            orders_skipped += skipped
            log_lines.append("📦 %s : %s commande(s) importée(s), %s ignorée(s) (déjà présentes ou vides)."
                              % (sheet_name.strip(), created, skipped))

        summary = (
            "Import terminé.\n\n"
            "Total : %s devis créés / %s ignorés — %s commandes créées / %s ignorées.\n\n"
            % (quotes_created, quotes_skipped, orders_created, orders_skipped)
        )
        self.result_log = summary + "\n".join(log_lines)

        return {
            'type': 'ir.actions.act_window',
            'res_model': 'b2fa.import.wizard',
            'view_mode': 'form',
            'res_id': self.id,
            'target': 'new',
        }

    def _detect_activity(self, norm_sheet_name):
        for token, activity in ACTIVITY_SHEET_MAP:
            if token in norm_sheet_name:
                return activity
        return False

    def _models_for_activity(self, activity):
        """Asie utilise des modèles totalement indépendants de B2B/Fund
        Raising (pas de champ 'activity_type' partagé)."""
        if activity == 'asie':
            return self.env['b2fa.quote.asie'], self.env['b2fa.order.asie']
        return self.env['b2fa.quote'], self.env['b2fa.order']

    def _iter_rows(self, ws, header_row, col_map):
        for r in range(header_row + 1, ws.max_row + 1):
            row_vals = {}
            for c, key in col_map.items():
                row_vals[key] = ws.cell(row=r, column=c).value
            if self._row_is_empty(row_vals):
                continue
            yield row_vals

    def _import_quote_sheet(self, ws, header_row, col_map, activity, Quote):
        created = skipped = 0
        has_activity_field = 'activity_type' in Quote._fields
        for row_vals in self._iter_rows(ws, header_row, col_map):
            source_ref = _to_str(row_vals.get('name')) or _to_str(row_vals.get('order_ref'))
            client = _to_str(row_vals.get('client'))
            if source_ref:
                domain = [('source_ref', '=', source_ref)]
                if has_activity_field:
                    domain.append(('activity_type', '=', activity))
                existing = Quote.search(domain, limit=1)
                if existing:
                    skipped += 1
                    continue

            vals = {
                'client': client or 'Client non renseigné',
                'secteur': _to_str(row_vals.get('secteur')),
                'contact': _to_str(row_vals.get('contact')),
                'description': _to_str(row_vals.get('description')),
                'qty': _to_float(row_vals.get('qty')) or 1.0,
                'amount': _to_float(row_vals.get('amount')),
                'state': self._map_state(row_vals.get('state'), QUOTE_STATE_MAP) or 'en_cours',
                'date_devis': _to_date(row_vals.get('date_devis')),
                'date_relance_1': _to_date(row_vals.get('date_relance_1')),
                'date_relance_2': _to_date(row_vals.get('date_relance_2')),
                'date_relance_3': _to_date(row_vals.get('date_relance_3')),
                'reponse_client': _to_str(row_vals.get('reponse_client')),
                'motif_refus': _to_str(row_vals.get('motif_refus')),
                'proba_conversion': self._map_proba(row_vals.get('proba_conversion')),
                'next_action': _to_str(row_vals.get('next_action')),
                'notes': _to_str(row_vals.get('notes')),
                'source_ref': source_ref,
            }
            if has_activity_field:
                vals['activity_type'] = activity
            # Char fields with False default would raise if not allowed; keep as empty string.
            for k in ('secteur', 'contact', 'description', 'reponse_client', 'motif_refus',
                      'next_action', 'notes'):
                if vals[k] is False:
                    vals[k] = ''
            Quote.create(vals)
            created += 1
        return created, skipped

    def _import_order_sheet(self, ws, header_row, col_map, activity, Order, Quote):
        created = skipped = 0
        has_activity_field = 'activity_type' in Order._fields
        for row_vals in self._iter_rows(ws, header_row, col_map):
            source_ref = _to_str(row_vals.get('name')) or _to_str(row_vals.get('quote_ref'))
            client = _to_str(row_vals.get('client'))
            if source_ref:
                domain = [('source_ref', '=', source_ref)]
                if has_activity_field:
                    domain.append(('activity_type', '=', activity))
                existing = Order.search(domain, limit=1)
                if existing:
                    skipped += 1
                    continue

            quote_ref_text = _to_str(row_vals.get('quote_ref'))
            quote = False
            if quote_ref_text:
                quote_domain = [('source_ref', '=', quote_ref_text)]
                if 'activity_type' in Quote._fields:
                    quote_domain.append(('activity_type', '=', activity))
                quote = Quote.search(quote_domain, limit=1)

            vals = {
                'client': client or 'Client non renseigné',
                'contact': _to_str(row_vals.get('contact')),
                'description': _to_str(row_vals.get('description')),
                'qty': _to_float(row_vals.get('qty')) or 1.0,
                'amount_ht': _to_float(row_vals.get('amount_ht')),
                'cost': _to_float(row_vals.get('cost')),
                'deposit_received': _to_float(row_vals.get('deposit_received')),
                'state': self._map_state(row_vals.get('state'), ORDER_STATE_MAP) or 'confirmee',
                'production_source': self._map_production_source(row_vals.get('production_source')),
                'date_commande': _to_date(row_vals.get('date_commande')),
                'date_prod_prevue': _to_date(row_vals.get('date_prod_prevue')),
                'date_prod_reelle': _to_date(row_vals.get('date_prod_reelle')),
                'date_expedition': _to_date(row_vals.get('date_expedition')),
                'carrier': _to_str(row_vals.get('carrier')),
                'logistics_fees': _to_float(row_vals.get('logistics_fees')),
                'tracking_number': _to_str(row_vals.get('tracking_number')),
                'date_livraison_prevue': _to_date(row_vals.get('date_livraison_prevue')),
                'date_livraison_reelle': _to_date(row_vals.get('date_livraison_reelle')),
                'notes_incidents': _to_str(row_vals.get('notes_incidents')),
                'source_ref': source_ref,
                'quote_ref_text': quote_ref_text or '',
                'quote_id': quote.id if quote else False,
            }
            if has_activity_field:
                vals['activity_type'] = activity
            for k in ('contact', 'description', 'carrier', 'tracking_number', 'notes_incidents'):
                if vals[k] is False:
                    vals[k] = ''
            Order.create(vals)
            created += 1
        return created, skipped
